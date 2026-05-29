"""workspace 源数据文件（xlsx / csv）存储与读取。"""

from __future__ import annotations

import csv
import json
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from skill_package.workspace.paths import config_path, ensure_workspace, validate_db_alias, workspace_dir

SOURCE_FILES_DIRNAME = "source_files"
SOURCE_FILE_MAX_BYTES = 20 * 1024 * 1024
ALLOWED_SOURCE_FILE_EXT = frozenset({".csv", ".xlsx"})


def source_files_dir(db_alias: str) -> Path:
    alias = validate_db_alias(db_alias)
    path = workspace_dir(alias) / SOURCE_FILES_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _safe_stored_name(original: str) -> str:
    p = Path(str(original or "").strip())
    ext = p.suffix.lower()
    if ext not in ALLOWED_SOURCE_FILE_EXT:
        raise ValueError(f"不支持的文件类型: {ext or '(无扩展名)'}，仅允许 csv、xlsx")
    stem = re.sub(r"[^\w\-]+", "_", p.stem).strip("_")[:80] or "file"
    return f"{datetime.now():%Y%m%d_%H%M%S}_{stem}{ext}"


def _validate_rel_path(rel_path: str) -> str:
    rel = str(rel_path or "").strip().replace("\\", "/").lstrip("/")
    if not rel.startswith(f"{SOURCE_FILES_DIRNAME}/"):
        raise ValueError(f"非法 source_files 路径: {rel_path!r}")
    if ".." in rel.split("/"):
        raise ValueError(f"非法 source_files 路径: {rel_path!r}")
    name = Path(rel).name
    if Path(name).suffix.lower() not in ALLOWED_SOURCE_FILE_EXT:
        raise ValueError(f"不支持的文件类型: {name}")
    return rel


def resolve_source_file_path(db_alias: str, rel_path: str) -> Path:
    alias = validate_db_alias(db_alias)
    rel = _validate_rel_path(rel_path)
    path = (workspace_dir(alias) / rel).resolve()
    root = workspace_dir(alias).resolve()
    if root not in path.parents:
        raise ValueError("source_files 路径须位于 workspace 内")
    return path


def normalize_source_file_entry(raw: Any) -> dict[str, Any] | None:
    if not isinstance(raw, dict):
        return None
    rel = str(raw.get("path") or "").strip()
    if not rel:
        return None
    try:
        rel = _validate_rel_path(rel)
    except ValueError:
        return None
    ext = Path(rel).suffix.lower().lstrip(".")
    name = str(raw.get("name") or Path(rel).name).strip() or Path(rel).name
    return {
        "id": str(raw.get("id") or uuid.uuid4().hex[:12]),
        "name": name,
        "path": rel,
        "type": ext if ext in ("csv", "xlsx") else Path(rel).suffix.lower().lstrip("."),
        "size": int(raw.get("size") or 0),
        "uploaded_at": str(raw.get("uploaded_at") or _now_iso()),
    }


def normalize_source_files(raw: Any) -> list[dict[str, Any]]:
    if not raw:
        return []
    if not isinstance(raw, list):
        return []
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in raw:
        entry = normalize_source_file_entry(item)
        if not entry or entry["path"] in seen:
            continue
        seen.add(entry["path"])
        out.append(entry)
    return out


def _load_config_raw(db_alias: str) -> dict[str, Any]:
    path = config_path(db_alias)
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, dict) else {}


def _write_source_files_config(db_alias: str, entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    from skill_package.workspace.paths import touch_manifest

    alias = validate_db_alias(db_alias)
    ensure_workspace(alias)
    path = config_path(alias)
    data = _load_config_raw(alias)
    if not data:
        data = {"db_alias": alias}
    normalized = normalize_source_files(entries)
    data["source_files"] = normalized
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    touch_manifest(alias, source_files=normalized)
    return normalized


def list_source_files(db_alias: str, *, verify_disk: bool = True) -> list[dict[str, Any]]:
    alias = validate_db_alias(db_alias)
    entries = normalize_source_files(_load_config_raw(alias).get("source_files"))
    if not verify_disk:
        return entries
    out: list[dict[str, Any]] = []
    for entry in entries:
        path = resolve_source_file_path(alias, entry["path"])
        if not path.is_file():
            continue
        item = dict(entry)
        item["size"] = path.stat().st_size
        out.append(item)
    if len(out) != len(entries):
        _write_source_files_config(alias, out)
    return out


def upload_source_files(db_alias: str, files: list[tuple[str, bytes]]) -> dict[str, Any]:
    alias = validate_db_alias(db_alias)
    if not files:
        raise ValueError("未选择文件")
    ensure_workspace(alias)
    dest_dir = source_files_dir(alias)
    existing = list_source_files(alias, verify_disk=True)
    by_path = {e["path"]: e for e in existing}
    uploaded: list[dict[str, Any]] = []

    for original_name, content in files:
        if not content:
            raise ValueError(f"文件为空: {original_name}")
        if len(content) > SOURCE_FILE_MAX_BYTES:
            raise ValueError(
                f"文件 {original_name} 超过大小限制 ({SOURCE_FILE_MAX_BYTES // (1024 * 1024)}MB)"
            )
        stored = _safe_stored_name(original_name)
        rel = f"{SOURCE_FILES_DIRNAME}/{stored}"
        path = dest_dir / stored
        path.write_bytes(content)
        entry = {
            "id": uuid.uuid4().hex[:12],
            "name": Path(original_name).name,
            "path": rel,
            "type": Path(stored).suffix.lower().lstrip("."),
            "size": len(content),
            "uploaded_at": _now_iso(),
        }
        by_path[rel] = entry
        uploaded.append(entry)

    merged = sorted(by_path.values(), key=lambda x: x.get("uploaded_at", ""), reverse=True)
    saved = _write_source_files_config(alias, merged)
    return {"uploaded": uploaded, "source_files": saved}


def delete_source_file(db_alias: str, rel_path: str) -> dict[str, Any]:
    alias = validate_db_alias(db_alias)
    rel = _validate_rel_path(rel_path)
    path = resolve_source_file_path(alias, rel)
    if path.is_file():
        path.unlink()
    remaining = [e for e in list_source_files(alias, verify_disk=False) if e["path"] != rel]
    saved = _write_source_files_config(alias, remaining)
    return {"deleted": rel, "source_files": saved}


def _decode_csv_bytes(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _preview_csv(data: bytes, *, max_rows: int) -> dict[str, Any]:
    text = _decode_csv_bytes(data)
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return {"columns": [], "row_count": 0, "sample_rows": []}
    columns = [str(c) for c in rows[0]]
    body = rows[1:]
    sample: list[dict[str, str]] = []
    for row in body[:max_rows]:
        sample.append({columns[i]: str(row[i]) if i < len(row) else "" for i in range(len(columns))})
    return {"columns": columns, "row_count": len(body), "sample_rows": sample}


def _preview_xlsx(data: bytes, *, sheet: str | None, max_rows: int) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("读取 xlsx 需要 openpyxl，请执行: pip install openpyxl") from e

    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    if not sheet_names:
        wb.close()
        return {"sheets": [], "columns": [], "row_count": 0, "sample_rows": []}
    active = sheet if sheet and sheet in sheet_names else sheet_names[0]
    ws = wb[active]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    columns = [str(c) if c is not None else "" for c in (header or [])]
    sample: list[dict[str, str]] = []
    row_count = 0
    for row in rows_iter:
        row_count += 1
        if len(sample) < max_rows:
            sample.append(
                {
                    columns[i]: "" if i >= len(row) or row[i] is None else str(row[i])
                    for i in range(len(columns))
                }
            )
    wb.close()
    return {
        "sheets": sheet_names,
        "sheet": active,
        "columns": columns,
        "row_count": row_count,
        "sample_rows": sample,
    }


def read_source_file_preview(
    db_alias: str,
    rel_path: str,
    *,
    sheet: str | None = None,
    max_rows: int = 20,
) -> dict[str, Any]:
    alias = validate_db_alias(db_alias)
    rel = _validate_rel_path(rel_path)
    path = resolve_source_file_path(alias, rel)
    if not path.is_file():
        raise FileNotFoundError(f"源文件不存在: {rel}")
    data = path.read_bytes()
    ext = path.suffix.lower()
    meta = {
        "db_alias": alias,
        "path": rel,
        "name": path.name,
        "type": ext.lstrip("."),
        "size": len(data),
    }
    if ext == ".csv":
        preview = _preview_csv(data, max_rows=max(1, min(max_rows, 100)))
    elif ext == ".xlsx":
        preview = _preview_xlsx(data, sheet=sheet, max_rows=max(1, min(max_rows, 100)))
    else:
        raise ValueError(f"不支持的文件类型: {ext}")
    return {**meta, **preview, "ok": True}
